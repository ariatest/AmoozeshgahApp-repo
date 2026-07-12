from __future__ import annotations

from acasmart.data.repos.reports_repo import get_all_student_terms_with_financials
from acasmart.data.repos.payments_repo import get_payments_with_notes_by_terms
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QHeaderView,
    QHBoxLayout, QLineEdit, QComboBox, QPushButton, QFileDialog
)
from acasmart.data.repos.classes_repo import fetch_classes
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor

from acasmart.core.utils import format_currency_with_unit
from acasmart.ui.widgets.shamsi_date_picker import ShamsiDatePicker
from acasmart.ui.widgets.theme_manager import ThemeManager
from acasmart.ui.widgets.base_secondary_window import BaseSecondaryWindow

import jdatetime
import openpyxl

# رنگِ مازاد (زرد) — هماهنگ با گزارش پرداخت‌ها
_EXTRA_YELLOW = QColor("#FFD54F")

# ستون‌های جدول
_HEADERS = [
    "نام هنرجو", "کلاس", "ساز", "استاد", "شروع ترم", "پایان ترم",
    "شهریه", "پرداخت شهریه", "مازاد", "بدهی", "توضیحات پرداخت‌ها", "وضعیت مالی"
]
_COL_EXTRA = 8   # مازاد
_COL_NOTES = 10  # توضیحات پرداخت‌ها
_COL_STATUS = 11


class FinancialReportWindow(BaseSecondaryWindow):
    def __init__(self, return_target: QWidget | None = None):
        super().__init__("گزارش مالی هنرجویان", return_target)
        self.setGeometry(300, 200, 1400, 500)
        self.all_data = []          # وضعیتِ ترم‌ها (شهریه/بدهی/وضعیت — مادام‌العمر)
        self.payments_by_term = {}  # {term_id: [(payment_date, payment_type, amount, description)]}
        self.filtered_data = []
        self.build_ui()

    def build_ui(self):
        layout = self.content_layout()
        layout.setSpacing(10)

        title = QLabel("📊 گزارش مالی ترم‌های هنرجویان")
        title.setProperty("sectionTitle", True)
        layout.addWidget(title)

        self.summary_label = QLabel("")
        self.summary_label.setProperty("caption", True)
        layout.addWidget(self.summary_label)

        # --- فیلترها (زنده؛ بدون دکمهٔ «اعمال فیلتر») ---
        filter_layout = QHBoxLayout()
        self.input_student_name = QLineEdit()
        self.input_student_name.setPlaceholderText("نام هنرجو")

        self.combo_class = QComboBox()
        self.combo_class.addItem("همه کلاس‌ها", None)
        for cid, cname, *_ in fetch_classes():
            self.combo_class.addItem(cname, cid)

        self.combo_status = QComboBox()
        self.combo_status.addItems(["همه وضعیت‌ها", "تسویه", "بدهکار"])

        # فیلترِ وضعیتِ ترم — نمایشِ ترم‌های فعال، تمام‌شده یا همه
        self.combo_term_status = QComboBox()
        self.combo_term_status.addItems(["همه ترم‌ها", "ترم فعال", "ترم تمام‌شده"])

        btn_clear = QPushButton("پاکسازی فیلتر")
        btn_clear.setProperty("variant", "secondary")
        btn_clear.clicked.connect(self.reset_filters)

        btn_export = QPushButton("📤 خروجی اکسل")
        btn_export.setProperty("variant", "ghost")
        btn_export.clicked.connect(self.export_to_excel)

        filter_layout.addWidget(self.input_student_name)
        filter_layout.addWidget(self.combo_class)
        filter_layout.addWidget(self.combo_status)
        filter_layout.addWidget(self.combo_term_status)
        filter_layout.addWidget(btn_clear)
        filter_layout.addWidget(btn_export)

        # بازهٔ تاریخِ پرداخت — مبالغِ «پرداخت» بر اساسِ همین بازه محاسبه می‌شوند
        self.date_from_picker = ShamsiDatePicker(": پرداخت از تاریخ")
        self.date_to_picker = ShamsiDatePicker(": تا تاریخ")
        # بدون تاریخِ پیش‌فرض؛ کاربر خودش بازهٔ دلخواه را انتخاب می‌کند
        self.date_from_picker.clear()
        self.date_to_picker.clear()
        filter_layout.addWidget(self.date_from_picker)
        filter_layout.addWidget(self.date_to_picker)

        layout.addLayout(filter_layout)

        # --- جدول ---
        self.table = QTableWidget()
        self.table.setColumnCount(len(_HEADERS))
        self.table.setHorizontalHeaderLabels(_HEADERS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # ستونِ توضیحات را محدود کن تا متنِ بلند بقیهٔ ستون‌ها را فشرده نکند؛
        # متن در سلول کوتاه می‌شود (…) و کاملش در tooltip می‌ماند. کاربر می‌تواند پهنایش را تغییر دهد.
        self.table.horizontalHeader().setSectionResizeMode(_COL_NOTES, QHeaderView.Interactive)
        self.table.setColumnWidth(_COL_NOTES, 240)
        self.table.setTextElideMode(Qt.ElideRight)
        self.table.setWordWrap(False)
        layout.addWidget(self.table)
        self.table.setSortingEnabled(True)
        self.showMaximized()

        for w in (
            title, self.summary_label, self.input_student_name,
            self.combo_class, self.combo_status, self.combo_term_status,
            btn_clear, btn_export, self.date_from_picker, self.date_to_picker, self.table,
        ):
            try:
                ThemeManager.repolish(w)
            except Exception:
                pass

        self.load_data()
        self._wire_live_filters()

    def _wire_live_filters(self):
        """فیلترِ زنده: هر تغییرِ ورودی بلافاصله اعمال می‌شود (بدون دکمهٔ «اعمال فیلتر»)."""
        apply = lambda *_: self.apply_filters()
        self.input_student_name.textChanged.connect(apply)
        for combo in (self.combo_class, self.combo_status, self.combo_term_status):
            combo.currentIndexChanged.connect(apply)
        self.date_from_picker.dateChanged.connect(apply)
        self.date_to_picker.dateChanged.connect(apply)

    def load_data(self):
        self.all_data = get_all_student_terms_with_financials()
        self.payments_by_term = get_payments_with_notes_by_terms([r['term_id'] for r in self.all_data])
        self.apply_filters()

    def _date_window(self):
        """بازهٔ تاریخِ پرداخت (شمسی)؛ هر کران ممکن است None باشد یعنی بدونِ قید."""
        frm = self.date_from_picker.selected_shamsi if self.date_from_picker.has_date() else None
        to = self.date_to_picker.selected_shamsi if self.date_to_picker.has_date() else None
        return frm, to

    def _payments_in_window(self, term_id, frm, to):
        """پرداخت‌های ترم که تاریخِ آن‌ها در بازهٔ انتخابی است (شمسی؛ مقایسهٔ رشته‌ای معتبر است)."""
        out = []
        for pdate, ptype, amount, desc in self.payments_by_term.get(term_id, []):
            if frm is not None and pdate < frm:
                continue
            if to is not None and pdate > to:
                continue
            out.append((pdate, ptype, amount, desc))
        return out

    def _notes_text(self, payments):
        """توضیحاتِ پرداخت‌های داده‌شده (شهریه و مازاد)."""
        label = {"tuition": "شهریه", "extra": "مازاد"}
        parts = []
        for _pdate, ptype, amount, desc in payments:
            line = f"{label.get(ptype, ptype)}: {format_currency_with_unit(amount)}"
            if desc:
                line += f" — {desc}"
            parts.append(line)
        return " | ".join(parts)

    def populate_table(self, data, window_active=False):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(data))
        total_tuition = total_paid = total_extra = total_debt = 0

        for i, row in enumerate(data):
            self.table.setItem(i, 0, QTableWidgetItem(row['student_name']))
            self.table.setItem(i, 1, QTableWidgetItem(row['class_name']))
            self.table.setItem(i, 2, QTableWidgetItem(row['instrument']))
            self.table.setItem(i, 3, QTableWidgetItem(row['teacher_name']))
            self.table.setItem(i, 4, QTableWidgetItem(row['start_date']))
            self.table.setItem(i, 5, QTableWidgetItem(row['end_date'] or "—"))
            self.table.setItem(i, 6, QTableWidgetItem(format_currency_with_unit(row['tuition'])))
            # پرداختِ شهریه — محدود به بازهٔ تاریخِ انتخاب‌شده
            self.table.setItem(i, 7, QTableWidgetItem(format_currency_with_unit(row['paid_tuition_win'])))

            # مازاد (زرد) — محدود به بازهٔ تاریخِ انتخاب‌شده
            paid_extra = row['paid_extra_win']
            item_extra = QTableWidgetItem(format_currency_with_unit(paid_extra))
            item_extra.setBackground(_EXTRA_YELLOW)
            item_extra.setForeground(QColor("#5C3D00"))
            self.table.setItem(i, _COL_EXTRA, item_extra)

            # بدهی و وضعیت مالی — وضعیتِ کلیِ ترم (مادام‌العمر)، مستقل از بازهٔ تاریخ
            self.table.setItem(i, 9, QTableWidgetItem(format_currency_with_unit(row['debt'])))

            notes = row['notes']
            item_notes = QTableWidgetItem(notes)
            item_notes.setToolTip(notes)
            self.table.setItem(i, _COL_NOTES, item_notes)

            item_status = QTableWidgetItem(row['status'])
            if row['status'] == "تسویه":
                item_status.setBackground(QColor("#81C784"))
                item_status.setForeground(QColor("#0B3D17"))
            elif row['status'] == "بدهکار":
                item_status.setBackground(QColor("#E57373"))
                item_status.setForeground(QColor("#5D1919"))
            item_status.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, _COL_STATUS, item_status)

            total_tuition += row['tuition']
            total_paid += row['paid_tuition_win']
            total_extra += paid_extra
            total_debt += row['debt']

        note = "   (پرداخت‌ها در بازهٔ تاریخِ انتخابی)" if window_active else ""
        self.summary_label.setText(
            f"تعداد ترم‌ها: {len(data)}   |   مجموع شهریه: {format_currency_with_unit(total_tuition)}"
            f"   |   مجموع پرداخت: {format_currency_with_unit(total_paid)}"
            f"   |   مجموع مازاد: {format_currency_with_unit(total_extra)}"
            f"   |   مجموع بدهی: {format_currency_with_unit(total_debt)}{note}"
        )
        self.table.setSortingEnabled(True)
        self.filtered_data = data

    def apply_filters(self):
        name_filter = self.input_student_name.text().strip()
        class_id = self.combo_class.currentData()
        status = self.combo_status.currentText()
        term_status = self.combo_term_status.currentText()
        # بازهٔ تاریخِ پرداخت — اختیاری؛ اگر فعال باشد، فقط درآمدِ همان بازه شمرده می‌شود
        frm, to = self._date_window()
        window_active = frm is not None or to is not None

        display = []
        for row in self.all_data:
            if name_filter and name_filter not in row['student_name']:
                continue
            if class_id and row['class_id'] != class_id:
                continue
            if status != "همه وضعیت‌ها" and row['status'] != status:
                continue
            # وضعیتِ ترم: فعال = بدون تاریخِ پایان، تمام‌شده = دارای تاریخِ پایان
            is_active = not row['end_date']
            if term_status == "ترم فعال" and not is_active:
                continue
            if term_status == "ترم تمام‌شده" and is_active:
                continue

            pays = self._payments_in_window(row['term_id'], frm, to)
            # حالتِ «درآمدِ دوره»: با بازهٔ فعال، فقط ترم‌هایی که در آن بازه پرداخت داشته‌اند نمایش داده می‌شوند
            if window_active and not pays:
                continue

            display.append({
                **row,
                'paid_tuition_win': sum(a for _d, t, a, _de in pays if t == 'tuition'),
                'paid_extra_win': sum(a for _d, t, a, _de in pays if t == 'extra'),
                'notes': self._notes_text(pays),
            })

        self.populate_table(display, window_active)

    def reset_filters(self):
        self.input_student_name.clear()
        self.combo_class.setCurrentIndex(0)
        self.combo_status.setCurrentIndex(0)
        self.combo_term_status.setCurrentIndex(0)
        self.date_from_picker.clear()
        self.date_to_picker.clear()
        self.apply_filters()

    def export_to_excel(self):
        today_shamsi = jdatetime.date.today().strftime("%Y-%m-%d")
        suggested_filename = f"گزارش مالی - {today_shamsi}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "ذخیره فایل اکسل", suggested_filename, "Excel Files (*.xlsx)")
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        for col, header in enumerate(_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, row_data in enumerate(self.filtered_data, start=2):
            ws.cell(row=row_idx, column=1, value=row_data['student_name'])
            ws.cell(row=row_idx, column=2, value=row_data['class_name'])
            ws.cell(row=row_idx, column=3, value=row_data['instrument'])
            ws.cell(row=row_idx, column=4, value=row_data['teacher_name'])
            ws.cell(row=row_idx, column=5, value=row_data['start_date'])
            ws.cell(row=row_idx, column=6, value=row_data['end_date'] or "—")
            ws.cell(row=row_idx, column=7, value=row_data['tuition'])
            ws.cell(row=row_idx, column=8, value=row_data['paid_tuition_win'])
            ws.cell(row=row_idx, column=9, value=row_data['paid_extra_win'])
            ws.cell(row=row_idx, column=10, value=row_data['debt'])
            ws.cell(row=row_idx, column=11, value=row_data['notes'])
            ws.cell(row=row_idx, column=12, value=row_data['status'])

        wb.save(file_path)
